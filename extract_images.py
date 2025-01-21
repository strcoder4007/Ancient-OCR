import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
from PIL import Image

wb = load_workbook('KaithiDataset.xlsx', keep_links=False)

sheets = ['Symbols', 'Characters', 'Words', 'Numbers']

image_path = 'extracted_images'
os.makedirs(image_path, exist_ok=True)

labels_data = []
image_counter = 1

for sheet_name in sheets:
    ws = wb[sheet_name]
    image_loader = SheetImageLoader(ws)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        label = row[0]
        for col_idx in range(2, 5):
            cell_ref = f'{chr(64 + col_idx)}{row_idx}' 
            if image_loader.image_in(cell_ref):
                try:
                    img = image_loader.get(cell_ref)
                    img_name = f'word_{image_counter}.png'
                    img_path = os.path.join(image_path, img_name)
                    img.save(img_path)
                    labels_data.append((img_name, label))
                    image_counter += 1
                except ValueError as e:
                    print(f"Error while processing image in {sheet_name} - {cell_ref}: {e}")
                    continue

labels_df = pd.DataFrame(labels_data, columns=['filename', 'words'])
labels_file = 'labels.csv'
labels_df.to_csv(labels_file, index=False)
print(f"Images and labels have been saved. Labels CSV: {labels_file}")